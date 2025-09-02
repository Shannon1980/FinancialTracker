"""
Comprehensive Data Import System for SEAS Financial Tracker
Handles Excel template import, validation, and population for employees and subcontractors
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
from typing import Dict, List, Tuple, Optional, Any
import io
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import tempfile
import os
from sample_data_generator import SampleDataGenerator


class DataImportSystem:
    """Handles Excel template import, validation, and population"""
    
    def __init__(self):
        self.template_columns = [
            'Name', 'Type', 'LCAT', 'Priced_Salary', 'Current_Salary', 
            'Hours_Per_Month', 'Department', 'Start_Date', 'Location', 
            'Manager', 'Skills', 'Contract_End_Date'
        ]
        
        self.validation_options = {
            'Type': ['Team', 'Subcontractor'],
            'LCAT': [
                'Project Manager', 'Senior Engineer', 'Full Stack Developer',
                'Data Engineer', 'Cloud Engineer', 'DevOps Engineer',
                'Business Analyst', 'UX/UI Designer', 'QA Engineer',
                'Technical Lead', 'Solution Architect', 'Consultant'
            ],
            'Department': [
                'Engineering', 'Data Engineering', 'DevOps', 'Product Management',
                'Design', 'Quality Assurance', 'Business Analysis', 'Management'
            ],
            'Location': ['Remote', 'On-site', 'Hybrid'],
            'Status': ['Active', 'Inactive', 'On Leave', 'Contract Ended']
        }
    
    def create_employee_template(self) -> bytes:
        """Create comprehensive Excel template for employee data"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Employee_Data sheet
        ws_data = wb.create_sheet("Employee_Data")
        
        # Define headers with monthly columns
        headers = self.template_columns.copy()
        
        # Add monthly columns (24 months from current date)
        current_date = datetime.now()
        monthly_columns = []
        for i in range(24):
            month_start = current_date + timedelta(days=30*i)
            month_end = month_start + timedelta(days=29)
            date_range = f"{month_start.strftime('%m/%d')}-{month_end.strftime('%m/%d/%y')}"
            monthly_columns.extend([f"Hours_{date_range}", f"Revenue_{date_range}"])
        
        headers.extend(monthly_columns)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add sample data
        sample_data = [
            {
                'Name': 'John Smith',
                'Type': 'Team',
                'LCAT': 'Senior Engineer',
                'Priced_Salary': 120000,
                'Current_Salary': 125000,
                'Hours_Per_Month': 173,
                'Department': 'Engineering',
                'Start_Date': '2024-01-15',
                'Location': 'Remote',
                'Manager': 'Jane Doe',
                'Skills': 'Python, React, AWS, Docker',
                'Contract_End_Date': ''
            },
            {
                'Name': 'Sarah Johnson',
                'Type': 'Subcontractor',
                'LCAT': 'Data Engineer',
                'Priced_Salary': 100000,
                'Current_Salary': 110000,
                'Hours_Per_Month': 120,
                'Department': 'Data Engineering',
                'Start_Date': '2024-03-01',
                'Location': 'Hybrid',
                'Manager': 'Mike Wilson',
                'Skills': 'SQL, Python, Spark, Azure',
                'Contract_End_Date': '2025-12-31'
            }
        ]
        
        # Add sample data rows
        for row_idx, data in enumerate(sample_data, 2):
            for col_idx, header in enumerate(self.template_columns, 1):
                value = data.get(header, '')
                ws_data.cell(row=row_idx, column=col_idx, value=value)
        
        # Add monthly data for sample rows
        for row_idx in range(2, 4):  # Sample rows
            for col_idx in range(len(self.template_columns) + 1, len(headers) + 1, 2):
                # Hours (full month for team, prorated for subcontractors)
                hours = 173 if ws_data.cell(row=row_idx, column=2).value == 'Team' else 120
                ws_data.cell(row=row_idx, column=col_idx, value=hours)
                
                # Revenue calculation
                salary = ws_data.cell(row=row_idx, column=5).value or 0
                hourly_rate = salary / (173 * 12) if salary else 0
                revenue = hours * hourly_rate * 1.2  # 20% overhead
                ws_data.cell(row=row_idx, column=col_idx + 1, value=round(revenue, 2))
        
        # Create Instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        
        instructions_data = [
            ['Field', 'Required', 'Description', 'Example'],
            ['Name', 'Yes', 'Full name of employee/contractor', 'John Smith'],
            ['Type', 'Yes', 'Team (internal) or Subcontractor (external)', 'Team'],
            ['LCAT', 'Yes', 'Labor Category/Job Title', 'Senior Engineer'],
            ['Priced_Salary', 'Yes', 'Budgeted salary for project', '120000'],
            ['Current_Salary', 'Yes', 'Actual current salary', '125000'],
            ['Hours_Per_Month', 'Yes', 'Expected hours per month', '173'],
            ['Department', 'No', 'Department or team', 'Engineering'],
            ['Start_Date', 'No', 'Start date (YYYY-MM-DD)', '2024-01-15'],
            ['Location', 'No', 'Work location type', 'Remote'],
            ['Manager', 'No', 'Direct manager name', 'Jane Doe'],
            ['Skills', 'No', 'Comma-separated skills', 'Python, React, AWS'],
            ['Contract_End_Date', 'No', 'End date for subcontractors', '2025-12-31']
        ]
        
        for row_idx, row_data in enumerate(instructions_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_instructions.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Create Validation_Options sheet
        ws_validation = wb.create_sheet("Validation_Options")
        
        validation_data = [
            ['Field', 'Valid Options'],
            ['Type', ', '.join(self.validation_options['Type'])],
            ['LCAT', ', '.join(self.validation_options['LCAT'])],
            ['Department', ', '.join(self.validation_options['Department'])],
            ['Location', ', '.join(self.validation_options['Location'])]
        ]
        
        for row_idx, row_data in enumerate(validation_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_validation.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Add data validation to Employee_Data sheet
        self._add_data_validation(ws_data, headers)
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _add_data_validation(self, worksheet, headers):
        """Add data validation to Excel worksheet"""
        # Type validation
        type_col = headers.index('Type') + 1
        type_validation = DataValidation(type="list", formula1=f'"{",".join(self.validation_options["Type"])}"')
        type_validation.add(f"{chr(64 + type_col)}2:{chr(64 + type_col)}1000")
        worksheet.add_data_validation(type_validation)
        
        # LCAT validation
        lcat_col = headers.index('LCAT') + 1
        lcat_validation = DataValidation(type="list", formula1=f'"{",".join(self.validation_options["LCAT"])}"')
        lcat_validation.add(f"{chr(64 + lcat_col)}2:{chr(64 + lcat_col)}1000")
        worksheet.add_data_validation(lcat_validation)
        
        # Department validation
        dept_col = headers.index('Department') + 1
        dept_validation = DataValidation(type="list", formula1=f'"{",".join(self.validation_options["Department"])}"')
        dept_validation.add(f"{chr(64 + dept_col)}2:{chr(64 + dept_col)}1000")
        worksheet.add_data_validation(dept_validation)
        
        # Location validation
        loc_col = headers.index('Location') + 1
        loc_validation = DataValidation(type="list", formula1=f'"{",".join(self.validation_options["Location"])}"')
        loc_validation.add(f"{chr(64 + loc_col)}2:{chr(64 + loc_col)}1000")
        worksheet.add_data_validation(loc_validation)
    
    def validate_employee_data(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Validate employee data according to business rules"""
        errors = []
        
        # Required fields validation
        required_fields = ['Name', 'Type', 'LCAT', 'Priced_Salary', 'Current_Salary', 'Hours_Per_Month']
        for field in required_fields:
            if field not in df.columns:
                errors.append(f"Missing required column: {field}")
                continue
            
            missing_values = df[field].isna().sum()
            if missing_values > 0:
                errors.append(f"Missing values in {field}: {missing_values} rows")
        
        # Type validation
        if 'Type' in df.columns:
            invalid_types = df[~df['Type'].isin(self.validation_options['Type'])]
            if not invalid_types.empty:
                errors.append(f"Invalid Type values: {invalid_types['Type'].unique().tolist()}")
        
        # LCAT validation
        if 'LCAT' in df.columns:
            invalid_lcats = df[~df['LCAT'].isin(self.validation_options['LCAT'])]
            if not invalid_lcats.empty:
                errors.append(f"Invalid LCAT values: {invalid_lcats['LCAT'].unique().tolist()}")
        
        # Salary validation
        if 'Priced_Salary' in df.columns:
            invalid_salaries = df[(df['Priced_Salary'] < 0) | (df['Priced_Salary'] > 500000)]
            if not invalid_salaries.empty:
                errors.append(f"Invalid Priced_Salary values (must be 0-500000): {len(invalid_salaries)} rows")
        
        if 'Current_Salary' in df.columns:
            invalid_salaries = df[(df['Current_Salary'] < 0) | (df['Current_Salary'] > 500000)]
            if not invalid_salaries.empty:
                errors.append(f"Invalid Current_Salary values (must be 0-500000): {len(invalid_salaries)} rows")
        
        # Hours validation
        if 'Hours_Per_Month' in df.columns:
            invalid_hours = df[(df['Hours_Per_Month'] < 0) | (df['Hours_Per_Month'] > 200)]
            if not invalid_hours.empty:
                errors.append(f"Invalid Hours_Per_Month values (must be 0-200): {len(invalid_hours)} rows")
        
        # Date validation
        if 'Start_Date' in df.columns:
            try:
                df['Start_Date'] = pd.to_datetime(df['Start_Date'], errors='coerce')
                invalid_dates = df[df['Start_Date'].isna()]
                if not invalid_dates.empty:
                    errors.append(f"Invalid Start_Date format: {len(invalid_dates)} rows")
            except:
                errors.append("Start_Date column contains invalid date formats")
        
        # Subcontractor specific validation
        if 'Type' in df.columns and 'Contract_End_Date' in df.columns:
            subcontractors = df[df['Type'] == 'Subcontractor']
            if not subcontractors.empty:
                missing_contract_dates = subcontractors[subcontractors['Contract_End_Date'].isna()]
                if not missing_contract_dates.empty:
                    errors.append(f"Subcontractors missing Contract_End_Date: {len(missing_contract_dates)} rows")
        
        return len(errors) == 0, errors
    
    def process_uploaded_data(self, uploaded_file) -> Tuple[pd.DataFrame, List[str]]:
        """Process uploaded Excel file and return validated DataFrame"""
        try:
            # Read the Excel file
            df = pd.read_excel(uploaded_file, sheet_name='Employee_Data')
            
            # Clean column names
            df.columns = df.columns.str.strip()
            
            # Add Type column if missing
            if 'Type' not in df.columns:
                df.insert(1, 'Type', 'Team')  # Default to Team
            
            # Validate data
            is_valid, errors = self.validate_employee_data(df)
            
            if not is_valid:
                return df, errors
            
            # Calculate monthly revenue if not provided
            df = self._calculate_monthly_revenue(df)
            
            return df, []
            
        except Exception as e:
            return pd.DataFrame(), [f"Error processing file: {str(e)}"]
    
    def _calculate_monthly_revenue(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calculate monthly revenue for employees"""
        # Find monthly columns
        monthly_columns = [col for col in df.columns if col.startswith('Hours_') or col.startswith('Revenue_')]
        
        for idx, row in df.iterrows():
            if pd.isna(row.get('Current_Salary')) or pd.isna(row.get('Hours_Per_Month')):
                continue
            
            # Calculate hourly rate
            hourly_rate = row['Current_Salary'] / (row['Hours_Per_Month'] * 12)
            
            # Apply markup for subcontractors
            markup = 1.2 if row.get('Type') == 'Subcontractor' else 1.0
            
            # Calculate revenue for each month
            for col in monthly_columns:
                if col.startswith('Revenue_'):
                    # Find corresponding hours column
                    hours_col = col.replace('Revenue_', 'Hours_')
                    if hours_col in df.columns:
                        hours = row.get(hours_col, 0)
                        if pd.isna(hours):
                            hours = row['Hours_Per_Month']
                        revenue = hours * hourly_rate * markup
                        df.at[idx, col] = round(revenue, 2)
        
        return df
    
    def generate_summary_report(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generate summary report for imported data"""
        if df.empty:
            return {"error": "No data to summarize"}
        
        summary = {
            "total_employees": len(df),
            "team_count": len(df[df['Type'] == 'Team']) if 'Type' in df.columns else 0,
            "subcontractor_count": len(df[df['Type'] == 'Subcontractor']) if 'Type' in df.columns else 0,
            "total_revenue_projection": 0,
            "average_salary": 0,
            "departments": {},
            "locations": {},
            "errors": []
        }
        
        # Calculate total revenue projection
        revenue_columns = [col for col in df.columns if col.startswith('Revenue_')]
        if revenue_columns:
            total_revenue = df[revenue_columns].sum().sum()
            summary["total_revenue_projection"] = round(total_revenue, 2)
        
        # Calculate average salary
        if 'Current_Salary' in df.columns:
            summary["average_salary"] = round(df['Current_Salary'].mean(), 2)
        
        # Department breakdown
        if 'Department' in df.columns:
            summary["departments"] = df['Department'].value_counts().to_dict()
        
        # Location breakdown
        if 'Location' in df.columns:
            summary["locations"] = df['Location'].value_counts().to_dict()
        
        return summary


def render_data_import_page():
    """Render the data import page with modern UI"""
    st.markdown("## 📥 Data Import & Management")
    
    # Initialize data import system
    if 'data_import_system' not in st.session_state:
        st.session_state.data_import_system = DataImportSystem()
    
    import_system = st.session_state.data_import_system
    
    # Create tabs for different import options
    tab1, tab2, tab3, tab4 = st.tabs(["📋 Download Template", "📤 Upload Data", "🎲 Generate Sample Data", "📊 Import Summary"])
    
    with tab1:
        st.markdown("### 📋 Employee Data Template")
        st.markdown("""
        Download the comprehensive Excel template to populate with your employee and subcontractor data.
        The template includes:
        - **Employee_Data**: Main data sheet with all required fields
        - **Instructions**: Detailed field descriptions and requirements
        - **Validation_Options**: Dropdown options for data validation
        """)
        
        # Template download button
        if st.button("📥 Download Excel Template", use_container_width=True):
            template_data = import_system.create_employee_template()
            st.download_button(
                label="📄 Download Template",
                data=template_data,
                file_name=f"Employee_Template_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("✅ Template downloaded successfully!")
        
        # Show template structure
        st.markdown("#### 📋 Template Structure")
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Required Fields:**")
            st.markdown("""
            - Name (Full name)
            - Type (Team/Subcontractor)
            - LCAT (Job title)
            - Priced_Salary (Budgeted)
            - Current_Salary (Actual)
            - Hours_Per_Month
            """)
        
        with col2:
            st.markdown("**Optional Fields:**")
            st.markdown("""
            - Department
            - Start_Date
            - Location
            - Manager
            - Skills
            - Contract_End_Date
            """)
    
    with tab2:
        st.markdown("### 📤 Upload Employee Data")
        
        # File upload
        uploaded_file = st.file_uploader(
            "Choose Excel file",
            type=['xlsx', 'xls'],
            help="Upload the populated Excel template with employee data"
        )
        
        if uploaded_file is not None:
            # Process uploaded file
            with st.spinner("Processing uploaded file..."):
                df, errors = import_system.process_uploaded_data(uploaded_file)
            
            if errors:
                st.error("❌ Validation Errors Found:")
                for error in errors:
                    st.error(f"• {error}")
            else:
                st.success("✅ File processed successfully!")
                
                # Show preview
                st.markdown("#### 📊 Data Preview")
                st.dataframe(df.head(10), use_container_width=True)
                
                # Store in session state
                st.session_state.employees = df
                
                # Show summary
                summary = import_system.generate_summary_report(df)
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Employees", summary["total_employees"])
                with col2:
                    st.metric("Team Members", summary["team_count"])
                with col3:
                    st.metric("Subcontractors", summary["subcontractor_count"])
                with col4:
                    st.metric("Avg Salary", f"${summary['average_salary']:,.0f}")
                
                # Department breakdown
                if summary["departments"]:
                    st.markdown("#### 🏢 Department Breakdown")
                    dept_df = pd.DataFrame(list(summary["departments"].items()), 
                                         columns=['Department', 'Count'])
                    st.bar_chart(dept_df.set_index('Department'))
    
    with tab3:
        st.markdown("### 🎲 Generate Sample Data")
        st.markdown("""
        Generate realistic sample data for testing and demonstration purposes.
        This will create a mix of team members and subcontractors with realistic salaries,
        skills, and monthly revenue projections.
        """)
        
        col1, col2 = st.columns(2)
        
        with col1:
            team_count = st.number_input("Number of Team Members", min_value=1, max_value=20, value=10)
        
        with col2:
            subcontractor_count = st.number_input("Number of Subcontractors", min_value=1, max_value=15, value=5)
        
        if st.button("🎲 Generate Sample Data", use_container_width=True):
            with st.spinner("Generating sample data..."):
                generator = SampleDataGenerator()
                df = generator.generate_complete_dataset(team_count, subcontractor_count)
                
                # Store in session state
                st.session_state.employees = df
                
                st.success(f"✅ Generated {len(df)} employees successfully!")
                
                # Show preview
                st.markdown("#### 📊 Sample Data Preview")
                st.dataframe(df.head(10), use_container_width=True)
                
                # Show summary
                summary = import_system.generate_summary_report(df)
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Employees", summary["total_employees"])
                with col2:
                    st.metric("Team Members", summary["team_count"])
                with col3:
                    st.metric("Subcontractors", summary["subcontractor_count"])
                with col4:
                    st.metric("Avg Salary", f"${summary['average_salary']:,.0f}")
        
        # Download sample data
        if 'employees' in st.session_state and not st.session_state.employees.empty:
            st.markdown("#### 📥 Download Sample Data")
            col1, col2 = st.columns(2)
            
            with col1:
                csv = st.session_state.employees.to_csv(index=False)
                st.download_button(
                    label="📊 Download as CSV",
                    data=csv,
                    file_name=f"sample_employee_data_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
            
            with col2:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    st.session_state.employees.to_excel(writer, sheet_name='Employee_Data', index=False)
                output.seek(0)
                st.download_button(
                    label="📋 Download as Excel",
                    data=output.getvalue(),
                    file_name=f"sample_employee_data_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    with tab4:
        st.markdown("### 📊 Import Summary")
        
        if 'employees' in st.session_state and not st.session_state.employees.empty:
            df = st.session_state.employees
            summary = import_system.generate_summary_report(df)
            
            # Summary metrics
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 📈 Financial Summary")
                st.metric("Total Revenue Projection", f"${summary['total_revenue_projection']:,.2f}")
                st.metric("Average Salary", f"${summary['average_salary']:,.2f}")
            
            with col2:
                st.markdown("#### 👥 Team Summary")
                st.metric("Total Employees", summary["total_employees"])
                st.metric("Team vs Subcontractors", f"{summary['team_count']} / {summary['subcontractor_count']}")
            
            # Export options
            st.markdown("#### 📤 Export Options")
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("📊 Export to CSV", use_container_width=True):
                    csv = df.to_csv(index=False)
                    st.download_button(
                        label="Download CSV",
                        data=csv,
                        file_name=f"employee_data_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv"
                    )
            
            with col2:
                if st.button("📋 Export to Excel", use_container_width=True):
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Employee_Data', index=False)
                    output.seek(0)
                    st.download_button(
                        label="Download Excel",
                        data=output.getvalue(),
                        file_name=f"employee_data_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.info("📝 No data imported yet. Use the Upload Data tab to import employee information.")
